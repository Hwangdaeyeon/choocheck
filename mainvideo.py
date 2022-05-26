from fileinput import filename
from keras.applications.mobilenet_v2 import preprocess_input
from tensorflow.python.keras.models import load_model
import numpy as np
import cv2, dlib
import face_function as ff
from PIL import ImageFont, ImageDraw, Image
from openpyxl import Workbook
import datetime

if __name__ == "__main__":
    facenet = cv2.dnn.readNet('models/deploy.prototxt', 'models/res10_300x300_ssd_iter_140000.caffemodel')
    model = load_model('8LBMI2.h5')

    # face encoding
    known_face_encodings, known_face_names = ff.face_library()
    fontpath = "fonts/gulim.ttc"
    font = ImageFont.truetype(fontpath, 30)
    # 실시간 웹캠 읽기
    cap = cv2.VideoCapture(0)
    i = 0

    img_frame = 5
    rst = 'unknown'

    # excel 활성화
    write_wb = Workbook()
    write_ws = write_wb.active
    cell_num = 1
    member = []

    while cap.isOpened():
        ret, img = cap.read()
        if not ret:
            break

        h, w = img.shape[:2]

        ori = img.copy()

        blob = cv2.dnn.blobFromImage(img, scalefactor=1., size=(405, 405), mean=(104., 177., 123.))
        facenet.setInput(blob)
        dets = facenet.forward()


        for i in range(dets.shape[2]):
            confidence = dets[0, 0, i, 2]
            if confidence < 0.5:
                continue

            x1 = int(dets[0, 0, i, 3] * w)
            y1 = int(dets[0, 0, i, 4] * h)
            x2 = int(dets[0, 0, i, 5] * w)
            y2 = int(dets[0, 0, i, 6] * h)

            face = img[y1:y2, x1:x2]
            face = face/256

            if (x2 >= w or y2 >= h):
                continue
            if (x1<=0 or y1<=0):
                continue

            face_input = cv2.resize(face,(200, 200))
            face_input = np.expand_dims(face_input, axis=0)
            face_input = np.array(face_input)

            modelpredict = model.predict(face_input)
            mask=modelpredict[0][0]
            nomask=modelpredict[0][1]

            # 좌표값
            top_left=np.array([x1, (y1+(y2-y1)/2)]).astype(int)
            bottom_right=np.array([x2+3, y2+22])
            center_xy=np.array([x1+(x2-x1)/2, y1+(3*(y2-y1)/4)]).astype(int)

            if mask > nomask:
                color = (0, 255, 0)
                label = 'Mask %d%%' % (mask * 100)
                

            else:
                color = (0, 0, 255)
                label = 'No Mask %d%%' % (nomask * 100)
                #frequency = 2500  # Set Frequency To 2500 Hertz
                #duration = 1000  # Set Duration To 1000 ms == 1 second
                #winsound.Beep(frequency, duration)


            cv2.rectangle(img, pt1=(x1, y1), pt2=(x2, y2), thickness=1, color=color, lineType=cv2.LINE_AA)
            cv2.putText(img, "Q : exit", org=(0, 20), fontFace=cv2.FONT_HERSHEY_SIMPLEX, fontScale=0.8,
                        color=(255,255,255), thickness=2, lineType=cv2.LINE_AA)
            cv2.putText(img, "A : registration", org=(0, 40), fontFace=cv2.FONT_HERSHEY_SIMPLEX, fontScale=0.8,
                        color=(255,255,255), thickness=2, lineType=cv2.LINE_AA)


        if ret:
            try:
                if (img_frame % 10) == 0 :
                    if mask > nomask :
                        compare_img = ff.img_overlay(top_left, bottom_right, img)
                        rst = ff.prt_result(compare_img, known_face_encodings, known_face_names)
                    else :
                        rst = ff.prt_result(img, known_face_encodings, known_face_names)

                if (rst != 'unknown') and (rst not in member) :
                    write_ws.cell(cell_num, 1, cell_num)
                    write_ws.cell(cell_num, 2, rst)
                    write_ws.cell(cell_num, 3, datetime.datetime.today())
                    cell_num = cell_num+1
                    member.append(rst)
                    print(member)
                else :
                    pass

                #print(img_frame)
                img_frame = img_frame+1
                img_pil = Image.fromarray(img)
                draw = ImageDraw.Draw(img_pil)
                draw.text((x1, y2 + 15), rst, font=font, fill=(255,255,255))
                img = np.array(img_pil)

                cv2.imshow('choolcheck',img)

            except:
                cv2.imshow('choolcheck',img)

            key = cv2.waitKey(1) & 0xFF

            if key == ord('a'):
                name = input("이름 : ")
                name_jpg = name + '.jpg'
                
                if mask > nomask :
                    ff.save_img(top_left, bottom_right, img, name)
                else :
                    cv2.imwrite('knowns_img/'+name+'.jpg', img)
                try:
                    known_face_encodings, known_face_names = ff.put_library(known_face_encodings, known_face_names, name_jpg)
                except:
                    print("등록 오류")

            if key == ord('q'):
                cap.release()
                cv2.destroyAllWindows
                write_wb.save("C:/choolcheck/test.xlsx")
                break
        else:
            break

    cap.release()
    cv2.destroyAllWindows
